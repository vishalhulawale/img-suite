"""XLSX converter — high-fidelity PDF table to Excel conversion.

Features:
  - Intelligent cell type detection (numbers, dates, currency, percentages)
  - Exact column widths and row heights
  - Cell formatting: fonts, colors, borders, alignment
  - Merged cell handling
  - Header row freeze
  - Multiple sheets (one per page or one per table)
"""

from __future__ import annotations

import logging
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
    numbers as xl_numbers,
)
from openpyxl.utils import get_column_letter

from app.pdf_converter.config import PipelineConfig
from app.pdf_converter.models import (
    BorderStyle, Color, Document, TableCell, TableElement,
    TextBlock, Page,
)

logger = logging.getLogger(__name__)

# Approximate column width: 1 Excel unit ≈ 7 pixels ≈ 5.25 pts
PTS_TO_EXCEL_WIDTH = 1.0 / 5.25
# Row height: 1 Excel unit = 1 pt
PTS_TO_EXCEL_HEIGHT = 1.0


def convert_to_xlsx(doc: Document, output_path: str, config: PipelineConfig) -> str:
    """Convert tables from the document model to an XLSX file."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    if config.xlsx.one_sheet_per_table:
        _export_one_sheet_per_table(wb, doc, config)
    elif config.xlsx.one_sheet_per_page:
        _export_one_sheet_per_page(wb, doc, config)
    else:
        _export_all_tables_one_sheet(wb, doc, config)

    # Ensure at least one sheet
    if not wb.sheetnames:
        ws = wb.create_sheet("Sheet1")
        ws["A1"] = "(No tables found in PDF)"

    wb.save(output_path)
    logger.info("XLSX saved to: %s", output_path)
    return output_path


# ---------------------------------------------------------------------------
# Export strategies
# ---------------------------------------------------------------------------

def _export_one_sheet_per_table(wb: Workbook, doc: Document, config: PipelineConfig):
    """Each detected table gets its own worksheet."""
    table_idx = 0
    for page in doc.pages:
        for table in page.tables:
            table_idx += 1
            name = _safe_sheet_name(f"Table {table_idx}")
            ws = wb.create_sheet(name)
            _write_table_to_sheet(ws, table, 1, 1, config)


def _export_one_sheet_per_page(wb: Workbook, doc: Document, config: PipelineConfig):
    """Each page gets a worksheet; multiple tables stack vertically."""
    for page_idx, page in enumerate(doc.pages):
        tables = page.tables
        if not tables:
            continue

        name = _safe_sheet_name(f"Page {page_idx + 1}")
        ws = wb.create_sheet(name)

        current_row = 1
        for table in tables:
            _write_table_to_sheet(ws, table, current_row, 1, config)
            current_row += table.num_rows + 2  # gap between tables


def _export_all_tables_one_sheet(wb: Workbook, doc: Document, config: PipelineConfig):
    """All tables on a single worksheet, stacked vertically."""
    ws = wb.create_sheet("Tables")
    current_row = 1

    for page in doc.pages:
        for table in page.tables:
            _write_table_to_sheet(ws, table, current_row, 1, config)
            current_row += table.num_rows + 2


# ---------------------------------------------------------------------------
# Core table writer
# ---------------------------------------------------------------------------

def _write_table_to_sheet(
    ws,
    table: TableElement,
    start_row: int,
    start_col: int,
    config: PipelineConfig,
):
    """Write a TableElement to an openpyxl worksheet at given position."""
    if table.num_rows == 0 or table.num_cols == 0:
        return

    # Set column widths
    for col_idx, width_pts in enumerate(table.col_widths):
        col_letter = get_column_letter(start_col + col_idx)
        excel_width = width_pts * PTS_TO_EXCEL_WIDTH
        if config.xlsx.auto_column_width:
            excel_width = max(8, min(config.xlsx.max_column_width, excel_width))
        ws.column_dimensions[col_letter].width = excel_width

    # Set row heights
    for row_idx, height_pts in enumerate(table.row_heights):
        ws.row_dimensions[start_row + row_idx].height = height_pts * PTS_TO_EXCEL_HEIGHT

    # Track merged cells
    merged: set[tuple[int, int]] = set()

    for cell in table.cells:
        xl_row = start_row + cell.row_index
        xl_col = start_col + cell.col_index

        if (cell.row_index, cell.col_index) in merged:
            continue

        # Handle merged cells
        if cell.row_span > 1 or cell.col_span > 1:
            end_row = xl_row + cell.row_span - 1
            end_col = xl_col + cell.col_span - 1
            try:
                ws.merge_cells(
                    start_row=xl_row, start_column=xl_col,
                    end_row=end_row, end_column=end_col,
                )
            except Exception:
                pass

            for r in range(cell.row_index, cell.row_index + cell.row_span):
                for c in range(cell.col_index, cell.col_index + cell.col_span):
                    if (r, c) != (cell.row_index, cell.col_index):
                        merged.add((r, c))

        xl_cell = ws.cell(row=xl_row, column=xl_col)

        # Cell value with type detection
        raw_text = cell.text.strip()
        value, number_format = _detect_cell_type(raw_text, config)
        xl_cell.value = value
        if number_format:
            xl_cell.number_format = number_format

        # Font formatting
        xl_cell.font = _make_excel_font(cell)

        # Alignment
        xl_cell.alignment = _make_excel_alignment(cell)

        # Background
        if cell.background_color:
            hex_color = cell.background_color.to_hex().lstrip("#")
            xl_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

        # Borders
        xl_cell.border = _make_excel_border(cell)

    # Freeze header row
    if config.xlsx.freeze_header_row and table.has_header_row:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col)


# ---------------------------------------------------------------------------
# Cell type detection
# ---------------------------------------------------------------------------

_CURRENCY_PATTERN = re.compile(
    r"^[£$€¥₹₽₩₦]?\s*-?\s*[\d,]+\.?\d*\s*[£$€¥₹₽₩₦]?$"
)
_PERCENT_PATTERN = re.compile(r"^-?\s*[\d,]+\.?\d*\s*%$")
_NUMBER_PATTERN = re.compile(r"^-?\s*[\d,]+\.?\d*$")
_DATE_PATTERNS = [
    (re.compile(r"^\d{1,2}/\d{1,2}/\d{2,4}$"), "%m/%d/%Y"),
    (re.compile(r"^\d{1,2}-\d{1,2}-\d{2,4}$"), "%m-%d-%Y"),
    (re.compile(r"^\d{4}-\d{2}-\d{2}$"), "%Y-%m-%d"),
    (re.compile(r"^\d{1,2}\s+\w+\s+\d{4}$"), "%d %B %Y"),
    (re.compile(r"^\w+\s+\d{1,2},?\s+\d{4}$"), "%B %d, %Y"),
]
_CURRENCY_SYMBOLS = {"$", "£", "€", "¥", "₹", "₽", "₩", "₦"}


def _detect_cell_type(
    text: str,
    config: PipelineConfig,
) -> tuple[str | float | datetime | None, str | None]:
    """Detect the type of a cell value and return (typed_value, number_format)."""
    if not text:
        return (None, None)

    # Percentage
    if config.xlsx.detect_number_formats and _PERCENT_PATTERN.match(text):
        try:
            num = float(text.replace("%", "").replace(",", "").strip())
            return (num / 100, "0.00%")
        except ValueError:
            pass

    # Currency
    if config.xlsx.detect_currency and _CURRENCY_PATTERN.match(text):
        clean = text
        symbol = ""
        for s in _CURRENCY_SYMBOLS:
            if s in clean:
                symbol = s
                clean = clean.replace(s, "")
                break

        clean = clean.replace(",", "").strip()
        try:
            num = float(clean)
            fmt_map = {
                "$": "$#,##0.00",
                "£": "£#,##0.00",
                "€": "€#,##0.00",
                "¥": "¥#,##0",
                "₹": "₹#,##0.00",
            }
            return (num, fmt_map.get(symbol, "#,##0.00"))
        except ValueError:
            pass

    # Date
    if config.xlsx.detect_dates:
        for pattern, fmt in _DATE_PATTERNS:
            if pattern.match(text):
                try:
                    dt = datetime.strptime(text.strip(","), fmt)
                    return (dt, "YYYY-MM-DD")
                except ValueError:
                    try:
                        fmt2 = fmt.replace("%Y", "%y")
                        dt = datetime.strptime(text.strip(","), fmt2)
                        return (dt, "YYYY-MM-DD")
                    except ValueError:
                        pass

    # Plain number
    if config.xlsx.detect_number_formats and _NUMBER_PATTERN.match(text):
        clean = text.replace(",", "").strip()
        try:
            num = float(clean)
            if num == int(num) and "." not in text:
                return (int(num), "#,##0")
            return (num, "#,##0.00")
        except ValueError:
            pass

    # Default: string
    return (text, None)


# ---------------------------------------------------------------------------
# Excel formatting helpers
# ---------------------------------------------------------------------------

def _make_excel_font(cell: TableCell) -> Font:
    """Create openpyxl Font from cell content formatting."""
    if cell.content and cell.content[0].lines and cell.content[0].lines[0].spans:
        span = cell.content[0].lines[0].spans[0]
        fi = span.font
        return Font(
            name=fi.name,
            size=fi.size,
            bold=fi.bold,
            italic=fi.italic,
            underline="single" if fi.underline else None,
            strike=fi.strikethrough,
            color=fi.color.to_hex().lstrip("#") if fi.color else "000000",
        )
    return Font(name="Calibri", size=11)


def _make_excel_alignment(cell: TableCell) -> Alignment:
    """Create openpyxl Alignment from cell properties."""
    from app.pdf_converter.models import TextAlignment

    h_align = "left"
    if cell.content:
        align = cell.content[0].alignment
        align_map = {
            TextAlignment.LEFT: "left",
            TextAlignment.CENTER: "center",
            TextAlignment.RIGHT: "right",
            TextAlignment.JUSTIFY: "justify",
        }
        h_align = align_map.get(align, "left")

    v_map = {"top": "top", "center": "center", "bottom": "bottom"}
    v_align = v_map.get(cell.vertical_alignment, "top")

    return Alignment(horizontal=h_align, vertical=v_align, wrap_text=True)


def _make_excel_border(cell: TableCell) -> Border:
    """Create openpyxl Border from cell border properties."""
    def make_side(border) -> Side:
        if border.style == BorderStyle.NONE:
            return Side(style=None)
        style_map = {
            BorderStyle.SOLID: "thin" if border.width <= 1 else "medium" if border.width <= 2 else "thick",
            BorderStyle.DASHED: "dashed",
            BorderStyle.DOTTED: "dotted",
            BorderStyle.DOUBLE: "double",
        }
        return Side(
            style=style_map.get(border.style, "thin"),
            color=border.color.to_hex().lstrip("#"),
        )

    return Border(
        top=make_side(cell.border_top),
        bottom=make_side(cell.border_bottom),
        left=make_side(cell.border_left),
        right=make_side(cell.border_right),
    )


def _safe_sheet_name(name: str) -> str:
    """Sanitize sheet name for Excel (max 31 chars, no special chars)."""
    for char in r"[]:*?/\\":
        name = name.replace(char, "")
    return name[:31]
